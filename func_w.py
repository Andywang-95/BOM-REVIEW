from copy import copy
from openpyxl import load_workbook
from openpyxl.styles import Protection
import datetime
import os

# # 輸出需要review的BOM的位置
# def BOM_file():
#     path = input('BOM FILE資料夾目錄：\n')
#     with os.scandir(path) as entries:
#         for entry in entries:
#             if 'BOM_TipTop_PTC' in entry.name:
#                 print('\n檔案名稱： '+ entry.name)
#                 return path.replace('\\','/')+'/'+entry.name
#     print('查無相關資料，請重新輸入\n\n')
#     return BOM_file()

# 確認maintain是否有被打開
def check_file(path):
    with os.scandir(path) as entries:
        for entry in entries:
            if '~$maintain' in entry.name:
                return '請關閉【maintain.xlsx】後再執行！\n-------------------------------------------------------------------------------'
    
# 載入工作簿與工作表
def load(path,name = None):
    wb = load_workbook(path)
    if name == None:
        ws = wb.active
    else:
        ws = wb[name]
    return wb,ws

# 更具料號進行maintain各工作表的匹配
def match(pn):
    electronic_1 = ['10DK','10DP','10DS','10DW','10DZ',
                '10GL','10HP','10IF','10IT','10LT',
                '10TA','11IF','11IT','11TA','11TC',
                '11TS','11TT','11WC','11WP','11WR',
                '10TC','10DE','10TT']
    electronic_2 = ['10CP','10CT','10DL','10DR',
                '10FF','10FP','10LB','10LC','10LF',
                '10LI','10LN','10OC','10OD','10XH',
                '10XT','11BL','11DL','11DR','11FP',
                '11LC','11LF','11XC','11XF','11XR']
    electronic_3 = ['10RC','10RH','10RN','10RS','10CE',
                '10CG','10CL','10CM','10CN','10CO',
                '10UA','10WW','10WR','10WP','10WA',
                '11BB','11CE','11CL','11CO','11RH']
    mechanism_list = ['10AC','10NH','10NR','10SA','10SB',
                    '10SC','10SL','10SM','10SR','10ST',
                    '11AC','11NH','11NR','11SA','11SC',
                    '11SI','11SM','11SR','12AC','12AI',
                    '12KR','10KS']

    maintain_dict = {
        '電子料(1)':electronic_1,
        '電子料(2)':electronic_2,
        '電子料(R,C)':electronic_3,
        '機構料件':mechanism_list
    }
    for name in maintain_dict:
        if pn[:4] in maintain_dict[name]:
            return name
    return 'Others'

# 將對應料號相關資料寫入maintain各個資料表中
def to_maintain(wb,d):
    count = {
    '電子料(1)':0,
    '電子料(2)':0,
    '電子料(R,C)':0,
    '機構料件':0,
    'Others':0
    }  
    for i in d:
        name = match(i)
        ws_maintain = wb[name]
        if i not in  [p.value for p in ws_maintain['A']]:
            ws_maintain.append([i,d[i][0].value,d[i][1].value,d[i][2].value,datetime.date.today()])
            ws_maintain['A'][-1].protection = Protection(locked=False)
            ws_maintain['D'][-1].protection = Protection(locked=False)
            ws_maintain['D'][-1].fill = copy(d[i][2].fill)
            ws_maintain['D'][-1].font = copy(d[i][2].font)
            ws_maintain['E'][-1].protection = Protection(locked=False)
            ws_maintain.protection.enable()
            count[name] += 1
    return count

# 轉換資料為字典形態(鍵為料號[A列],值為儲存格位置[C列])
def to_dict(ws):
    pn = [pn.value.strip() for pn in ws['A']][1:]
    d1 = ws['B'][1:]
    d2 = ws['C'][1:]
    ce = ws['D'][1:]
    dict_i = {}
    for p,d1,d2,ce in zip(pn,d1,d2,ce):
        dict_i[p] = (d1,d2,ce)
    return dict_i

# 更新mapping時，選擇要上傳更新的maintain工作表
# def sheet_name():
#     sheet_name = input('''請輸入要更新的工作表對應代碼：
#                         1 : 機構料件
#                         2 : Jack
#                         3 : Laney
#                         4 : Andy
#                         5 : Others
# 請輸入代碼→''')
#     if sheet_name == '1':
#         return '機構料件'
#     elif sheet_name == '2':
#         return 'Jack'
#     elif sheet_name == '3':
#         return 'Laney'
#     elif sheet_name == '4':
#         return 'Andy'
#     elif sheet_name == '5':
#         return 'Others'
#     else:
#         return '輸入錯誤'

# # 確認開始執行程式
# def start():
#     while True:
#         answer = input('請輸入 start 開始執行 \n') 
#         if answer != 'start':
#             os._exit(0)
#         else:
#             break

# # 確認退出程式
# def exit():
#     while True:
#         answer = input('退出請按 Enter \n') 
#         if answer == '':
#             os._exit(0)
    
